#!/usr/bin/env python3
"""Generate a job application tracking spreadsheet in XLSX format.

Usage:
    python job_tracker.py --output tracker.xlsx [--applications applications.json]

Optional input JSON (pre-populate with existing applications):
[
    {
        "company": "TechCorp",
        "position": "Senior Engineer",
        "url": "https://techcorp.com/jobs/123",
        "date_applied": "2026-02-20",
        "status": "Applied",
        "contact": "Jane Smith, Recruiter",
        "salary_range": "$180K-220K",
        "notes": "Referral from Bob"
    }
]

Status values: Researching, Applied, Phone Screen, Technical, Onsite, Offer, Accepted, Rejected, Withdrawn
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADERS = [
    ('Company', 22),
    ('Position', 25),
    ('URL', 35),
    ('Date Applied', 14),
    ('Status', 16),
    ('Contact', 25),
    ('Salary Range', 16),
    ('Next Step', 22),
    ('Follow-Up Date', 14),
    ('Interview Date', 14),
    ('Response Date', 14),
    ('Excitement (1-5)', 14),
    ('Notes', 40),
]

STATUS_OPTIONS = [
    'Researching', 'Applied', 'Phone Screen', 'Technical Interview',
    'Onsite Interview', 'Offer', 'Accepted', 'Rejected', 'Withdrawn', 'Ghosted'
]

HEADER_FILL = PatternFill('solid', fgColor='1A1A2E')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
ALT_FILL = PatternFill('solid', fgColor='F5F5F5')
BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

STATUS_COLORS = {
    'Researching': 'E8EAF6',
    'Applied': 'FFF3E0',
    'Phone Screen': 'E3F2FD',
    'Technical Interview': 'E8F5E9',
    'Onsite Interview': 'F3E5F5',
    'Offer': 'C8E6C9',
    'Accepted': '81C784',
    'Rejected': 'FFCDD2',
    'Withdrawn': 'CFD8DC',
    'Ghosted': 'F5F5F5',
}


def create_tracker(output_path, applications=None):
    wb = Workbook()

    # ── Applications Sheet ──
    ws = wb.active
    ws.title = 'Applications'
    ws.freeze_panes = 'A2'

    # Headers
    for col_idx, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Status dropdown validation
    status_list = ','.join(STATUS_OPTIONS)
    dv = DataValidation(type='list', formula1=f'"{status_list}"', allow_blank=True)
    dv.error = 'Please select a valid status'
    dv.prompt = 'Choose application status'
    ws.add_data_validation(dv)
    dv.add(f'E2:E200')

    # Excitement validation (1-5)
    ev = DataValidation(type='whole', operator='between', formula1='1', formula2='5', allow_blank=True)
    ws.add_data_validation(ev)
    ev.add('L2:L200')

    # Pre-populate data
    if applications:
        field_map = {
            'company': 1, 'position': 2, 'url': 3, 'date_applied': 4,
            'status': 5, 'contact': 6, 'salary_range': 7, 'next_step': 8,
            'follow_up_date': 9, 'interview_date': 10, 'response_date': 11,
            'excitement': 12, 'notes': 13
        }
        for row_idx, app in enumerate(applications, 2):
            for field, col in field_map.items():
                value = app.get(field, '')
                if value:
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    if row_idx % 2 == 0:
                        cell.fill = ALT_FILL

                    # Color-code status
                    if field == 'status' and value in STATUS_COLORS:
                        cell.fill = PatternFill('solid', fgColor=STATUS_COLORS[value])

    # ── Dashboard Sheet ──
    dash = wb.create_sheet('Dashboard')
    dash_headers = ['Metric', 'Value']
    for col_idx, header in enumerate(dash_headers, 1):
        cell = dash.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    dash.column_dimensions['A'].width = 30
    dash.column_dimensions['B'].width = 20

    metrics = [
        ('Total Applications', '=COUNTA(Applications!A2:A200)'),
        ('Applied', '=COUNTIF(Applications!E2:E200,"Applied")'),
        ('Phone Screens', '=COUNTIF(Applications!E2:E200,"Phone Screen")'),
        ('Technical Interviews', '=COUNTIF(Applications!E2:E200,"Technical Interview")'),
        ('Onsite Interviews', '=COUNTIF(Applications!E2:E200,"Onsite Interview")'),
        ('Offers', '=COUNTIF(Applications!E2:E200,"Offer")'),
        ('Accepted', '=COUNTIF(Applications!E2:E200,"Accepted")'),
        ('Rejected', '=COUNTIF(Applications!E2:E200,"Rejected")'),
        ('Withdrawn', '=COUNTIF(Applications!E2:E200,"Withdrawn")'),
        ('Ghosted', '=COUNTIF(Applications!E2:E200,"Ghosted")'),
        ('', ''),
        ('Response Rate', '=IF(B2>0,(B2-B10-B11)/B2,"N/A")'),
        ('Interview Conversion', '=IF(B3>0,(B4+B5+B6)/B3,"N/A")'),
        ('Offer Rate', '=IF(B2>0,B7/B2,"N/A")'),
        ('Avg Excitement', '=IF(COUNTA(Applications!L2:L200)>0,AVERAGE(Applications!L2:L200),"N/A")'),
    ]

    for row_idx, (label, formula) in enumerate(metrics, 2):
        cell_a = dash.cell(row=row_idx, column=1, value=label)
        cell_b = dash.cell(row=row_idx, column=2, value=formula)
        cell_a.font = Font(name='Calibri', size=10, bold=True)
        cell_b.font = DATA_FONT
        cell_a.border = BORDER
        cell_b.border = BORDER

    # ── Networking Sheet ──
    net = wb.create_sheet('Networking')
    net_headers = [
        ('Contact Name', 22), ('Company', 20), ('Title', 20),
        ('How Connected', 20), ('Last Contact', 14),
        ('Next Follow-Up', 14), ('Notes', 35)
    ]
    for col_idx, (header, width) in enumerate(net_headers, 1):
        cell = net.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        net.column_dimensions[get_column_letter(col_idx)].width = width

    net.freeze_panes = 'A2'

    wb.save(output_path)

    app_count = len(applications) if applications else 0
    return {
        "status": "success",
        "output": output_path,
        "sheets": ["Applications", "Dashboard", "Networking"],
        "applications_loaded": app_count,
        "features": [
            "Status dropdown with color coding",
            "Dashboard with conversion metrics",
            "Networking contact tracker",
            "Excitement rating (1-5)",
            "Follow-up date tracking"
        ]
    }


def main():
    parser = argparse.ArgumentParser(description='Generate job application tracker')
    parser.add_argument('--output', required=True, help='Output XLSX file path')
    parser.add_argument('--applications', help='Optional JSON file with existing applications')
    args = parser.parse_args()

    applications = None
    if args.applications:
        with open(args.applications, 'r') as f:
            applications = json.load(f)

    result = create_tracker(args.output, applications)
    print(json.dumps(result))


if __name__ == '__main__':
    main()
